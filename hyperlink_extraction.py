import openpyxl
import re

def extract_id_from_text(text):
    """Extract content inside parentheses like (GOV-1_10) from text"""
    if not text:
        return None
    match = re.search(r'\((.*?)\)', text)
    if not match:
        return None

    content = match.group(1)
    return content.split('.')[-1] 

def get_hyperlink_for_id(id_value, esrs_sheet):
    """Find hyperlink for given ID in ESRS sheet"""
    # Assuming:
    # - ID is in column A (index 0)
    # - Hyperlink is in column F (index 5)
    # - Headers are in row 2, data starts from row 3
    for row in esrs_sheet.iter_rows(min_row=3, values_only=False):
        esrs_id = row[0].value  # Column A (ID)
        name_cell = row[5]      # Column F (Name with hyperlink)
        
        if esrs_id == id_value:
            if name_cell.hyperlink:
                return name_cell.hyperlink.display  # Use .target for the actual URL
    return None

def add_hyperlinks_to_disclosure_requirements(file1_path, file2_path):
    # Load the first file (contains the disclosure requirements)
    wb1 = openpyxl.load_workbook(file1_path)
    
    # Load the second file (contains the ESRS 2 sheet)
    wb2 = openpyxl.load_workbook(file2_path, keep_vba=True)
    esrs_sheet = wb2['ESRS 2']  # Assuming the sheet name is 'ESRS 2'

    for sheet_name in wb1.sheetnames:
        sheet1 = wb1[sheet_name]
        
        # Find or create Hyperlink column
        headers = [cell.value for cell in sheet1[1]]
        if "Hyperlink" not in headers:
            hyperlink_col_idx = len(headers) + 1
            sheet1.cell(row=1, column=hyperlink_col_idx, value="Hyperlink")
        else:
            hyperlink_col_idx = headers.index("Hyperlink") + 1

        # Process each row (starting from row 2)
        for row_idx in range(2, sheet1.max_row + 1):
            disclosure_text = sheet1.cell(row=row_idx, column=3).value  # Assuming column B
            
            if disclosure_text:
                extracted_id = extract_id_from_text(disclosure_text)
                if extracted_id:
                    # First try ESRS 2 sheet
                    esrs_sheet = wb2['ESRS 2']
                    hyperlink = None
                    
                    if esrs_sheet:
                        hyperlink = get_hyperlink_for_id(extracted_id, esrs_sheet)
                    
                    # If not found in ESRS 2, search all other sheets
                    if not hyperlink:
                        for sheet in wb2.sheetnames:
                            if sheet != 'ESRS 2':  # Skip ESRS 2 as we already checked it
                                current_sheet = wb2[sheet]
                                hyperlink = get_hyperlink_for_id(extracted_id, current_sheet)
                                if hyperlink:
                                    break
                    sheet1.cell(row=row_idx, column=hyperlink_col_idx, value=hyperlink)
    
    # Save the modified first file
    wb1.save(file1_path)

# Example usage
file1_path = "/home/rajeshgupta/Git_Clones/esrs_gap_analysis/Download_ESRS GAP_Twilio (1).xlsx"
file2_path = "/home/rajeshgupta/Downloads/EFRAG IG 3 List of ESRS Data Points (1).xlsx"

add_hyperlinks_to_disclosure_requirements(file1_path, file2_path)